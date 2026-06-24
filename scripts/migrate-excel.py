from __future__ import annotations

import json
import re
import shutil
from collections import defaultdict
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = ROOT.parent

WORKBOOKS = [
    {
        "product_id": "ias-3120",
        "name": "IAS-3120",
        "workbook": SOURCE_ROOT / "Attachment 1-IAS-3120 Maintenance SOP.xlsx",
    },
    {
        "product_id": "ias-5100",
        "name": "IAS-5100",
        "workbook": SOURCE_ROOT / "Attachment 2-IAS-5100 Maintenance SOP.xlsx",
    },
]


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower()).strip("-")
    return slug or "sheet"


def ensure_unique_slug(slug: str, used: set[str]) -> str:
    candidate = slug
    index = 2
    while candidate in used:
        candidate = f"{slug}-{index}"
        index += 1
    used.add(candidate)
    return candidate


def json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def display_value(value: Any) -> str:
    value = json_value(value)
    if value is None:
        return ""
    return str(value)


def color_to_hex(color: Any) -> str | None:
    if color is None:
        return None
    color_type = getattr(color, "type", None)
    rgb = getattr(color, "rgb", None)
    if color_type == "rgb" and rgb:
        rgb = str(rgb)
        if len(rgb) == 8:
            rgb = rgb[2:]
        return f"#{rgb.upper()}"
    return None


def cell_style(cell: Any) -> dict[str, Any]:
    style: dict[str, Any] = {}

    fill = getattr(cell, "fill", None)
    fill_color = color_to_hex(getattr(fill, "fgColor", None))
    if fill_color and fill_color != "#000000":
        style["fill"] = fill_color

    font = getattr(cell, "font", None)
    font_data: dict[str, Any] = {}
    font_color = color_to_hex(getattr(font, "color", None))
    if getattr(font, "bold", False):
        font_data["bold"] = True
    if getattr(font, "italic", False):
        font_data["italic"] = True
    if font_color:
        font_data["color"] = font_color
    if getattr(font, "sz", None):
        font_data["size"] = float(font.sz)
    if font_data:
        style["font"] = font_data

    alignment = getattr(cell, "alignment", None)
    align_data: dict[str, Any] = {}
    for key in ("horizontal", "vertical", "wrap_text"):
        value = getattr(alignment, key, None)
        if value:
            align_data[key.replace("_", "-")] = value
    if align_data:
        style["alignment"] = align_data

    border = getattr(cell, "border", None)
    border_sides = []
    for key in ("left", "right", "top", "bottom"):
        side = getattr(border, key, None)
        if getattr(side, "style", None):
            border_sides.append(key)
    if border_sides:
        style["borders"] = border_sides

    if getattr(cell, "number_format", None) and cell.number_format != "General":
        style["numberFormat"] = cell.number_format

    return style


def merge_map(ws: Any) -> tuple[dict[tuple[int, int], dict[str, int]], set[tuple[int, int]], list[str]]:
    starts: dict[tuple[int, int], dict[str, int]] = {}
    covered: set[tuple[int, int]] = set()
    ranges: list[str] = []
    for merged in ws.merged_cells.ranges:
        ranges.append(str(merged))
        row_span = merged.max_row - merged.min_row + 1
        col_span = merged.max_col - merged.min_col + 1
        starts[(merged.min_row, merged.min_col)] = {"rowspan": row_span, "colspan": col_span}
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                if row == merged.min_row and col == merged.min_col:
                    continue
                covered.add((row, col))
    return starts, covered, ranges


def image_extension(image: Any) -> str:
    fmt = (getattr(image, "format", None) or "png").lower()
    if fmt == "jpeg":
        return "jpg"
    return fmt


def image_anchor_data(image: Any) -> dict[str, Any]:
    anchor = image.anchor
    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)
    data = {
        "from": {
            "row": getattr(start, "row", 0) + 1,
            "col": getattr(start, "col", 0) + 1,
            "rowOffset": getattr(start, "rowOff", 0),
            "colOffset": getattr(start, "colOff", 0),
        }
    }
    if end is not None:
        data["to"] = {
            "row": getattr(end, "row", 0) + 1,
            "col": getattr(end, "col", 0) + 1,
            "rowOffset": getattr(end, "rowOff", 0),
            "colOffset": getattr(end, "colOff", 0),
        }
    return data


def extract_sheet(ws: Any, product_id: str, sheet_slug: str) -> tuple[dict[str, Any], dict[str, Any]]:
    starts, covered, merge_ranges = merge_map(ws)
    image_dir = ROOT / "public" / "images" / "products" / product_id / "excel" / sheet_slug
    image_dir.mkdir(parents=True, exist_ok=True)

    row_heights = {
        str(index): dimension.height
        for index, dimension in ws.row_dimensions.items()
        if dimension.height is not None
    }
    column_widths = {
        letter: dimension.width
        for letter, dimension in ws.column_dimensions.items()
        if dimension.width is not None
    }

    rows = []
    non_empty = 0
    formula_count = 0
    hyperlink_count = 0
    comment_count = 0

    for row_index in range(1, ws.max_row + 1):
        cells = []
        for col_index in range(1, ws.max_column + 1):
            if (row_index, col_index) in covered:
                cells.append({"covered": True})
                continue

            cell = ws.cell(row=row_index, column=col_index)
            value = json_value(cell.value)
            text = display_value(cell.value)
            if cell.value is not None:
                non_empty += 1
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            if is_formula:
                formula_count += 1
            if cell.hyperlink:
                hyperlink_count += 1
            if cell.comment:
                comment_count += 1

            cell_data: dict[str, Any] = {
                "address": cell.coordinate,
                "row": row_index,
                "col": col_index,
                "column": get_column_letter(col_index),
                "value": value,
                "text": text,
            }

            if (row_index, col_index) in starts:
                cell_data["merge"] = starts[(row_index, col_index)]
            if is_formula:
                cell_data["formula"] = cell.value
            if cell.hyperlink:
                cell_data["hyperlink"] = cell.hyperlink.target
            if cell.comment:
                cell_data["comment"] = {
                    "author": cell.comment.author,
                    "text": cell.comment.text,
                }
            style = cell_style(cell)
            if style:
                cell_data["style"] = style

            cells.append(cell_data)
        rows.append({"index": row_index, "cells": cells})

    images = []
    for index, image in enumerate(getattr(ws, "_images", []), 1):
        ext = image_extension(image)
        filename = f"image-{index:03d}.{ext}"
        target = image_dir / filename
        target.write_bytes(image._data())
        images.append(
            {
                "id": f"image-{index:03d}",
                "file": f"./images/products/{product_id}/excel/{sheet_slug}/{filename}",
                "width": getattr(image, "width", None),
                "height": getattr(image, "height", None),
                "format": ext,
                "anchor": image_anchor_data(image),
            }
        )

    sheet_data = {
        "title": ws.title,
        "slug": sheet_slug,
        "dimensions": {
            "maxRow": ws.max_row,
            "maxColumn": ws.max_column,
            "rowHeights": row_heights,
            "columnWidths": column_widths,
        },
        "mergedCells": merge_ranges,
        "images": images,
        "rows": rows,
    }

    audit = {
        "title": ws.title,
        "slug": sheet_slug,
        "maxRow": ws.max_row,
        "maxColumn": ws.max_column,
        "nonEmptyCells": non_empty,
        "mergedCells": len(merge_ranges),
        "images": len(images),
        "formulas": formula_count,
        "hyperlinks": hyperlink_count,
        "comments": comment_count,
        "exportedRows": len(rows),
        "exportedCells": sum(len(row["cells"]) for row in rows),
        "status": "ok",
    }
    return sheet_data, audit


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def migrate_workbook(config: dict[str, Any]) -> dict[str, Any]:
    product_id = config["product_id"]
    workbook_path = config["workbook"]
    content_dir = ROOT / "public" / "content" / "products" / product_id / "excel"
    sheets_dir = content_dir / "sheets"

    sheets_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
    used_slugs: set[str] = set()
    manifest_sheets = []
    audits = []

    for index, ws in enumerate(wb.worksheets, 1):
        slug = ensure_unique_slug(slugify(ws.title), used_slugs)
        sheet_data, audit = extract_sheet(ws, product_id, slug)
        sheet_file = sheets_dir / f"{slug}.json"
        write_json(sheet_file, sheet_data)
        manifest_sheets.append(
            {
                "order": index,
                "title": ws.title,
                "slug": slug,
                "path": f"./public/content/products/{product_id}/excel/sheets/{slug}.json",
                "maxRow": audit["maxRow"],
                "maxColumn": audit["maxColumn"],
                "nonEmptyCells": audit["nonEmptyCells"],
                "mergedCells": audit["mergedCells"],
                "images": audit["images"],
            }
        )
        audits.append(audit)

    wb.close()

    manifest = {
        "productId": product_id,
        "productName": config["name"],
        "sourceWorkbook": workbook_path.name,
        "sheetCount": len(manifest_sheets),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sheets": manifest_sheets,
    }
    write_json(content_dir / "manifest.json", manifest)

    summary = {
        "productId": product_id,
        "sourceWorkbook": workbook_path.name,
        "sheetCount": len(audits),
        "totalRows": sum(item["maxRow"] for item in audits),
        "totalNonEmptyCells": sum(item["nonEmptyCells"] for item in audits),
        "totalMergedCells": sum(item["mergedCells"] for item in audits),
        "totalImages": sum(item["images"] for item in audits),
        "totalFormulas": sum(item["formulas"] for item in audits),
        "sheets": audits,
    }
    write_json(content_dir / "migration.audit.json", summary)
    return summary


def write_audit_markdown(summaries: list[dict[str, Any]]) -> None:
    lines = [
        "# Excel 内容迁移审核报告",
        "",
        "本报告由 `scripts/migrate-excel.py` 生成，用于核对 Excel 原始内容与网页数据文件的对应关系。",
        "",
    ]

    for summary in summaries:
        lines.extend(
            [
                f"## {summary['productId']}",
                "",
                f"- 原始文件：`{summary['sourceWorkbook']}`",
                f"- 工作表数量：{summary['sheetCount']}",
                f"- 总行数：{summary['totalRows']}",
                f"- 非空单元格：{summary['totalNonEmptyCells']}",
                f"- 合并区域：{summary['totalMergedCells']}",
                f"- 图片：{summary['totalImages']}",
                f"- 公式：{summary['totalFormulas']}",
                "",
                "| # | 工作表 | 行 | 列 | 非空单元格 | 合并区域 | 图片 | 状态 |",
                "|---:|---|---:|---:|---:|---:|---:|---|",
            ]
        )
        for index, sheet in enumerate(summary["sheets"], 1):
            lines.append(
                f"| {index} | {sheet['title']} | {sheet['maxRow']} | {sheet['maxColumn']} | "
                f"{sheet['nonEmptyCells']} | {sheet['mergedCells']} | {sheet['images']} | {sheet['status']} |"
            )
        lines.append("")

    (ROOT / "EXCEL_EXPORT_AUDIT.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    summaries = [migrate_workbook(config) for config in WORKBOOKS]
    write_audit_markdown(summaries)
    print(json.dumps({"status": "ok", "products": summaries}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
